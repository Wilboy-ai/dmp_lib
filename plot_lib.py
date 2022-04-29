import openpyxl
from pathlib import Path
import math
import matplotlib.pyplot as plt


def plot_T(x, y, z, set_title):
    fig = plt.figure()
    ax = plt.axes(projection='3d')
    ax.plot3D(x, y, z, 'red')
    ax.set_title(set_title)
    plt.show()

def plot_t(T, set_title):
    plt.plot(T, 'red')
    plt.title(set_title)
    plt.show()


def read_T(file_name, threshold=-1):
    # Read the active sheet:
    xlsx_file = Path('', file_name)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    x = []
    y = []
    z = []

    px = -100
    py = -100
    pz = -100


    for row in sheet.iter_rows(1, sheet.max_row):
        _x = row[0].value
        _y = row[1].value
        _z = row[2].value

        dist = math.sqrt(pow(_x - px,2) + pow(_y - py,2) + pow(_z - pz,2) )

        if dist >= threshold or threshold == -1:
            x.append(_x)
            y.append(_y)
            z.append(_z)

        px = _x
        py = _y
        pz = _z

    return x, y, z