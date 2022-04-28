#%matplotlib PyQt5
import urx

import openpyxl
from pathlib import Path

import math
import numpy as np
from numpy import diff

#matplotlib.use("Agg")
#%matplotlib inline
#import matplotlib
import matplotlib.pyplot as plt
from mpl_toolkits import mplot3d

alpha = 0.5
beta =0.5
tau = 0.1

orientation_fixed = (0.020, -3.172, 0.021)



def _plot_trajectory(x, y, z, set_title):
	fig = plt.figure()
	ax = plt.axes(projection='3d')
	ax.plot3D(x, y, z, 'gray')
	ax.set_title(set_title)
	plt.show()


def _read_T(file_name, threshold = 0.01):
	# Read the active sheet:
	xlsx_file = Path('', file_name)
	wb_obj = openpyxl.load_workbook(xlsx_file) 
	sheet = wb_obj.active

	x = []
	y = []
	z = []

	px = -100
	py = -100
	pz = -100

	for row in sheet.iter_rows(1, sheet.max_row):
		_x = row[0].value
		_y = row[1].value
		_z = row[2].value

		dist = math.sqrt( pow(_x - px,2) + pow(_y - py,2) + pow(_z - pz,2) )

		if dist >= threshold or threshold == -1:
			x.append(_x)
			y.append(_y)
			z.append(_z)

		px = _x
		py = _y
		pz = _z

	return x, y, z




def _calc_fd(T, y0, g):
	#g = T[len(T)-1]
	#y0 = T[0]

	dT = diff(T).tolist()
	dT.append(0)

	ddT	= diff(dT)
	ddT = ddT.tolist()
	ddT.append(0)

	fd = []
	for i in range(0, len(T)-1):
		f_target = ddT[i] - alpha * ( beta * (g - T[i]) - dT[i] )
		fd.append(f_target)

	#plt.plot(fd)
	#plt.ylabel('f_target')
	#plt.show()

	return fd

def _dmp_trajectory(L, T0, g, fd, label):

	T = T0
	dT = 0

	_dmp = []

	for i in range(0, L):
		ddT = alpha * ( beta * (g - T) - dT) + fd[i]
		dT = dT + tau * ddT
		T = T + tau * dT

		_dmp.append(T)

	#plt.plot(_dmp)
	#plt.ylabel(label)
	#plt.show()


	return _dmp


def _dmp(x, y, z):
	fdx = _calc_fd(x, x[0], x[len(x)-1])
	fdy = _calc_fd(y, y[0], y[len(y)-1])
	fdz = _calc_fd(z, z[0], z[len(z)-1])

	L = len(x)-1
	dmp_x = _dmp_trajectory(L, x[0], x[len(x)-1], fdx, 'x')
	dmp_y = _dmp_trajectory(L, y[0], y[len(y)-1], fdy, 'y')
	dmp_z = _dmp_trajectory(L, z[0], z[len(z)-1], fdz, 'z')

	return dmp_x, dmp_y, dmp_z


def _dmp_p(x, y, z, p0, pf):
	fdx = _calc_fd(x, p0[0], pf[0])
	fdy = _calc_fd(y, p0[1], pf[1])
	fdz = _calc_fd(z, p0[2], pf[2])
	#fdx = _calc_fd(x, x[0], x[len(x)-1])
	#fdy = _calc_fd(y, y[0], y[len(y)-1])
	#fdz = _calc_fd(z, z[0], z[len(z)-1])

	L = len(x)-1
	dmp_x = _dmp_trajectory(L, p0[0], pf[0], fdx, 'x')
	dmp_y = _dmp_trajectory(L, p0[1], pf[1], fdy, 'y')
	dmp_z = _dmp_trajectory(L, p0[2], pf[2], fdz, 'z')

	return dmp_x, dmp_y, dmp_z




def get_trajectory(file_name, threshold=0.01):
	x, y, z = _read_T(file_name, threshold)

	#_plot_trajectory(x, y, z, file_name)

	dmp_x, dmp_y, dmp_z = _dmp(x, y, z)

	fig = plt.figure()
	ax = plt.axes(projection='3d')
	ax.plot3D(x, y, z, 'red')
	ax.plot3D(dmp_x, dmp_y, dmp_z, 'blue')
	plt.show()

	print(len(dmp_x))

	dmp_x.reverse()
	dmp_y.reverse()
	dmp_z.reverse()


	T = []
	for i in range(0, len(dmp_x)-1):
		point = (dmp_x[i], dmp_y[i], dmp_z[i],  orientation_fixed(0), orientation_fixed(1), orientation_fixed(2))
		T.append(point)

	return T


def get_trajectory_point(file_name, p0, threshold=0.01):
	x, y, z = _read_T(file_name, threshold)

	print("trajectory points")
	print([x[0], y[0], z[0]])
	print([x[len(x)-1], y[len(y)-1], z[len(z)-1]])

	#_p0 = [x[len(x)-1], y[len(y)-1], z[len(z)-1]]
	#_pf = p0

	_p0 = p0 #[x[0], y[0], z[0]]
	#_p0 = [x[0], y[0], z[0]]
	_pf = [x[len(x)-1], y[len(y)-1], z[len(z)-1]]

	print("Target points")
	print(_p0)
	print(_pf)
	#_plot_trajectory(x, y, z, file_name)

	dmp_x, dmp_y, dmp_z = _dmp_p(x, y, z, _p0, _pf)
	
	dmp_x.reverse()
	dmp_y.reverse()
	dmp_z.reverse()

	fig = plt.figure()
	ax = plt.axes(projection='3d')
	ax.plot3D(x, y, z, 'red')
	ax.plot3D(dmp_x, dmp_y, dmp_z, 'blue')
	plt.show()

	T = []
	for i in range(0, len(dmp_x)-1):
		point = (dmp_x[i], dmp_y[i], dmp_z[i], orientation_fixed(0), orientation_fixed(1), orientation_fixed(2))
		T.append(point)

	return T


#T = get_trajectory('Position_testing2.xlsx', 0.0001)
T = get_trajectory_point('Position_testing_2.xlsx', [0.376,-0.176,0.395], 0.001)


#T = get_trajectory('T.xlsx', -1)
#T = get_trajectory_point('T.xlsx', [0.1,0.1,0.1], -1)

a = 0.1
v = 0.1

#rob = urx.Robot("192.168.1.100")
#rob.movels(T, a, v, 0.05, wait=True, threshold=0.001)
#rob.close()




	







